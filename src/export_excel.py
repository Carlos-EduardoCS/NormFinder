"""
export_excel.py
----------------

Responsável por exportar os resultados encontrados
para um arquivo Excel (.xlsx).

Autor: NormFinder
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExportExcel:

    def __init__(self):
        pass

    # ---------------------------------------------------------

    def exportar(self, resultados, arquivo_saida):

        """
        resultados -> lista retornada pelo scanner

        arquivo_saida -> caminho do arquivo .xlsx
        """

        wb = Workbook()
        ws = wb.active

        ws.title = "Resultados"

        # Cabeçalhos
        cabecalhos = [
            "Arquivo",
            "Caminho",
            "Norma",
            "Ocorrências",
            "Trecho"
        ]

        fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78",
            end_color="1F4E78"
        )

        fonte = Font(
            bold=True,
            color="FFFFFF"
        )

        for coluna, titulo in enumerate(cabecalhos, start=1):

            celula = ws.cell(
                row=1,
                column=coluna
            )

            celula.value = titulo
            celula.font = fonte
            celula.fill = fill
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        linha = 2

        for resultado in resultados:

            arquivo = resultado.get("arquivo", "")
            caminho = resultado.get("caminho", "")
            norma = resultado.get("norma", "")
            quantidade = resultado.get("quantidade", 0)

            trechos = resultado.get("trechos", [])

            if len(trechos) == 0:

                ws.cell(linha, 1).value = arquivo
                ws.cell(linha, 2).value = caminho
                ws.cell(linha, 3).value = norma
                ws.cell(linha, 4).value = quantidade
                ws.cell(linha, 5).value = ""

                linha += 1

            else:

                for trecho in trechos:

                    ws.cell(linha, 1).value = arquivo
                    ws.cell(linha, 2).value = caminho
                    ws.cell(linha, 3).value = norma
                    ws.cell(linha, 4).value = quantidade
                    ws.cell(linha, 5).value = trecho

                    linha += 1

        # Ajuste automático das colunas
        for coluna in ws.columns:

            maior = 0

            letra = get_column_letter(coluna[0].column)

            for celula in coluna:

                try:

                    tamanho = len(str(celula.value))

                    if tamanho > maior:
                        maior = tamanho

                except:
                    pass

            largura = min(maior + 3, 70)

            ws.column_dimensions[letra].width = largura

        # Quebra de linha no trecho
        for linha_excel in ws.iter_rows(min_row=2):

            linha_excel[4].alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

        # Congela o cabeçalho
        ws.freeze_panes = "A2"

        # Cria pasta caso não exista
        Path(arquivo_saida).parent.mkdir(
            parents=True,
            exist_ok=True
        )

        wb.save(arquivo_saida)

        return arquivo_saida


# ---------------------------------------------------------
# TESTE
# ---------------------------------------------------------

if __name__ == "__main__":

    resultados = [

        {
            "arquivo": "Manual.docx",

            "caminho": r"C:\Projetos\Manual.docx",

            "norma": "IEC 60335-2-15",

            "quantidade": 2,

            "trechos": [

                "Este produto atende à IEC 60335-2-15...",

                "Conforme IEC 60335-2-15 seção 19..."

            ]
        },

        {
            "arquivo": "Projeto.docx",

            "caminho": r"C:\Projetos\Projeto.docx",

            "norma": "IEC 60335-1",

            "quantidade": 1,

            "trechos": [

                "O equipamento está conforme IEC 60335-1."

            ]
        }

    ]

    exportador = ExportExcel()

    exportador.exportar(

        resultados,

        "Relatorio_Normas.xlsx"

    )

    print("Arquivo Excel criado com sucesso.")